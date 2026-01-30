import os
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, Response, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, date, timedelta
import cv2
import numpy as np
import json
import base64
import shutil
import io

# --- New Imports for MTCNN and InceptionResnetV1 ---
import torch
from torchvision import transforms
from PIL import Image
from facenet_pytorch import MTCNN, InceptionResnetV1

# --- Import for Text-to-Speech ---
import pyttsx3
import threading

# --- Import for Excel Export ---
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import openpyxl.utils

# --- Configuration ---
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
STATIC_FOLDER = os.path.join(BASE_DIR, 'static')
KNOWN_FACES_DIR = os.path.join(STATIC_FOLDER, 'known_faces')
INSTANCE_FOLDER_PATH = os.path.join(BASE_DIR, 'instance')
MODELS_DIR = os.path.join(BASE_DIR, 'models')

if not os.path.exists(INSTANCE_FOLDER_PATH):
    os.makedirs(INSTANCE_FOLDER_PATH)
if not os.path.exists(KNOWN_FACES_DIR):
    os.makedirs(KNOWN_FACES_DIR)
if not os.path.exists(MODELS_DIR):
    os.makedirs(MODELS_DIR)
    print(f"INFO: Models directory created at {MODELS_DIR}.")

# --- Device Configuration for PyTorch ---
device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
print(f"INFO: Running face recognition on device: {device}")

# --- App Initialization ---
app = Flask(__name__, instance_path=INSTANCE_FOLDER_PATH)
app.config['SECRET_KEY'] = 'your_very_secret_key_for_we_guide'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(app.instance_path, 'site.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['KNOWN_FACES_DIR'] = KNOWN_FACES_DIR

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'admin_login'
login_manager.login_message_category = 'info'

# --- New Model Placeholders & Gallery ---
mtcnn_detector = None
face_embedder_new = None

employee_gallery = {}
NEW_EMBEDDING_THRESHOLD = 0.75
EXPECTED_EMBEDDING_DIM = 512

# --- Context Processor ---
@app.context_processor
def inject_utilities():
    return {'now': datetime.utcnow(), 'json': json}

# --- Text-to-Speech Initialization ---
def initialize_tts():
    """Test TTS availability"""
    try:
        test_engine = pyttsx3.init()
        test_engine.stop()
        del test_engine
        print("INFO: Text-to-Speech is available and working.")
        return True
    except Exception as e:
        print(f"WARNING: Could not initialize TTS engine: {e}")
        return False

def speak_async(text):
    """Speak text in a separate thread to avoid blocking"""
    def speak_thread():
        try:
            # Create a new TTS engine instance for each speech request
            # This prevents issues with engine state and threading
            engine = pyttsx3.init()
            engine.setProperty('rate', 150)
            engine.setProperty('volume', 0.9)
            engine.say(text)
            engine.runAndWait()
            engine.stop()
            del engine
        except Exception as e:
            print(f"ERROR: TTS failed: {e}")
    
    thread = threading.Thread(target=speak_thread, daemon=True)
    thread.start()

# --- New Model Loading Function ---
def load_new_face_models():
    global mtcnn_detector, face_embedder_new
    print("INFO: Loading MTCNN for detection and InceptionResnetV1 for embedding...")
    try:
        mtcnn_detector = MTCNN(image_size=160, margin=14, min_face_size=40,
                               keep_all=True, post_process=True, device=device, select_largest=False)
        if mtcnn_detector is None:
            print("CRITICAL: MTCNN model could not be initialized.")
            return False
        print("INFO: MTCNN model loaded successfully.")

        face_embedder_new = InceptionResnetV1(pretrained='vggface2', device=device).eval()
        if face_embedder_new is None:
            print("CRITICAL: InceptionResnetV1 model could not be initialized.")
            return False
        print("INFO: InceptionResnetV1 (VGGFace2) embedder loaded successfully.")
        return True
    except Exception as e:
        print(f"CRITICAL: Error loading new models: {e}")
        mtcnn_detector = None
        face_embedder_new = None
        return False

# --- Database Models ---
class AdminUser(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)

class Employee(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.String(50), unique=True, nullable=False)
    name = db.Column(db.String(100), nullable=False)
    phone = db.Column(db.String(20))
    address = db.Column(db.String(200))
    dob = db.Column(db.Date)
    designation = db.Column(db.String(100), nullable=True)
    face_encodings = db.Column(db.Text, default='[]')
    registered_on = db.Column(db.DateTime, default=datetime.utcnow)
    attendances = db.relationship('AttendanceRecord', backref='employee_ref', lazy=True, cascade="all, delete-orphan")

class AttendanceRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employee.id'), nullable=False)
    timestamp_in = db.Column(db.DateTime, nullable=True)
    timestamp_out = db.Column(db.DateTime, nullable=True)
    date = db.Column(db.Date, default=date.today, nullable=False)
    status = db.Column(db.String(20), default='Present')

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(AdminUser, int(user_id))

# --- Face Processing Utilities ---
def detect_faces_and_get_aligned_tensors(cv2_frame):
    if mtcnn_detector is None:
        print("ERROR: MTCNN model not loaded for detection.")
        return [], None
    try:
        rgb_cv2_frame = cv2.cvtColor(cv2_frame, cv2.COLOR_BGR2RGB)
        pil_image = Image.fromarray(rgb_cv2_frame)
        batch_boxes, batch_probs, batch_landmarks = mtcnn_detector.detect(pil_image, landmarks=True)
        aligned_face_tensors = mtcnn_detector(pil_image)
        extracted_boxes = []
        if batch_boxes is not None:
            for i, box in enumerate(batch_boxes):
                if box is not None:
                    extracted_boxes.append(box.astype(int))
        if batch_boxes is None or aligned_face_tensors is None:
            return [], None
        return extracted_boxes, aligned_face_tensors
    except Exception as e:
        print(f"ERROR: Exception during MTCNN face detection/alignment: {e}")
        return [], None

def get_face_embedding_from_aligned_tensor(aligned_face_tensor_single):
    if face_embedder_new is None or aligned_face_tensor_single is None:
        print("ERROR: Face embedder model not loaded or no aligned face tensor provided.")
        return None
    try:
        face_tensor_batched = aligned_face_tensor_single.unsqueeze(0).to(device)
        with torch.no_grad():
            embedding = face_embedder_new(face_tensor_batched)
        return embedding.squeeze().cpu()
    except Exception as e:
        print(f"ERROR: Exception during new embedding extraction: {e}")
        return None

def load_known_face_embeddings_from_db_to_gallery():
    global employee_gallery
    employee_gallery = {}
    employees = Employee.query.all()
    loaded_embeddings_count = 0
    valid_embeddings_for_gallery = 0
    if not employees:
        print("INFO: No employees found in the database to load embeddings from.")
        return
    for emp in employees:
        if emp.face_encodings and emp.face_encodings != '[]':
            try:
                embeddings_list_of_lists = json.loads(emp.face_encodings)
                emp_embeddings_tensors = []
                for idx, emb_list in enumerate(embeddings_list_of_lists):
                    if len(emb_list) == EXPECTED_EMBEDDING_DIM:
                        emb_tensor = torch.tensor(emb_list, dtype=torch.float32, device='cpu')
                        emp_embeddings_tensors.append(emb_tensor)
                    else:
                        print(f"WARNING: Employee {emp.employee_id} (DB ID: {emp.id}), embedding index {idx} has incorrect dimension {len(emb_list)}. Expected {EXPECTED_EMBEDDING_DIM}. Skipping this embedding.")
                if emp_embeddings_tensors:
                    employee_gallery[emp.employee_id] = emp_embeddings_tensors
                    loaded_embeddings_count += len(embeddings_list_of_lists)
                    valid_embeddings_for_gallery += len(emp_embeddings_tensors)
            except json.JSONDecodeError:
                print(f"WARNING: Invalid JSON in face_encodings for employee {emp.employee_id} (ID: {emp.id}). Skipping.")
            except Exception as e:
                print(f"ERROR: Error loading embeddings for employee {emp.employee_id} (ID: {emp.id}) into gallery: {e}")
    if not employee_gallery:
        print(f"WARNING: Employee gallery is empty. Attempted to load {loaded_embeddings_count} embeddings, but 0 were valid for the gallery.")
    else:
        print(f"INFO: Attempted to load {loaded_embeddings_count} embeddings from DB. Added {valid_embeddings_for_gallery} valid embeddings for {len(employee_gallery)} employees into the gallery.")

def recognize_face_from_frame_new(cv2_frame):
    if not employee_gallery:
        return None, []
    face_boxes, aligned_face_tensors = detect_faces_and_get_aligned_tensors(cv2_frame)
    recognized_employee_id = None
    best_match_box_coords = []
    if aligned_face_tensors is not None and len(aligned_face_tensors) > 0:
        current_aligned_tensor = aligned_face_tensors[0]
        current_box = face_boxes[0] if face_boxes else []
        current_embedding = get_face_embedding_from_aligned_tensor(current_aligned_tensor)
        if current_embedding is not None:
            if current_embedding.shape[0] != EXPECTED_EMBEDDING_DIM:
                print(f"ERROR: Current embedding has unexpected dimension: {current_embedding.shape[0]}. Expected {EXPECTED_EMBEDDING_DIM}. Cannot compare.")
                return None, current_box
            max_similarity = -2.0
            candidate_employee_id = None
            current_embedding_gpu = current_embedding.to(device)
            for emp_id, known_embeddings_tensor_list in employee_gallery.items():
                for known_embedding_tensor_cpu in known_embeddings_tensor_list:
                    if known_embedding_tensor_cpu.shape[0] != EXPECTED_EMBEDDING_DIM:
                        continue
                    known_embedding_tensor_gpu = known_embedding_tensor_cpu.to(device)
                    sim = torch.nn.functional.cosine_similarity(
                        current_embedding_gpu.unsqueeze(0),
                        known_embedding_tensor_gpu.unsqueeze(0)
                    )
                    similarity_value = sim.item()
                    if similarity_value > max_similarity:
                        max_similarity = similarity_value
                        candidate_employee_id = emp_id
            if max_similarity >= NEW_EMBEDDING_THRESHOLD:
                recognized_employee_id = candidate_employee_id
                best_match_box_coords = current_box
    return recognized_employee_id, best_match_box_coords

# --- Routes ---
@app.route('/')
def index():
    if mtcnn_detector is None or face_embedder_new is None:
        flash("Face recognition models are not loaded. Please check server logs.", "danger")
    return render_template('index.html', title="Mark Attendance")

@app.route('/mark_attendance', methods=['POST'])
def mark_attendance_route():
    if mtcnn_detector is None or face_embedder_new is None:
        return jsonify({'status': 'error', 'message': 'Face recognition system not ready.'}), 503

    data = request.get_json()
    if not data or 'image_data' not in data:
        return jsonify({'status': 'error', 'message': 'No image data received.'}), 400

    auto_mode = data.get('auto_mode', False)
    recognize_only = data.get('recognize_only', False)  # New flag
    
    image_data_url = data['image_data']
    try:
        header, encoded = image_data_url.split(",", 1)
        image_data_bytes = base64.b64decode(encoded)
        nparr = np.frombuffer(image_data_bytes, np.uint8)
        frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    except Exception as e:
        print(f"ERROR decoding image: {e}")
        return jsonify({'status': 'error', 'message': f'Error decoding image: {str(e)}'}), 400

    if frame is None:
        return jsonify({'status': 'error', 'message': 'Could not decode image.'}), 400

    employee_id_recognized, _ = recognize_face_from_frame_new(frame)

    if employee_id_recognized:
        employee = Employee.query.filter_by(employee_id=employee_id_recognized).first()
        if employee:
            # If recognize_only flag is set, just return the employee_id without marking attendance
            if recognize_only:
                return jsonify({
                    'status': 'info',
                    'message': f'Recognized: {employee.name}',
                    'employee_id': employee_id_recognized
                })
            
            # Otherwise, proceed with marking attendance
            today = date.today()
            now_utc = datetime.utcnow()

            latest_record = AttendanceRecord.query.filter_by(
                employee_id=employee.id, date=today
            ).order_by(AttendanceRecord.id.desc()).first()

            message = ""
            status_code = 'info'
            speech_text = ""

            if latest_record and latest_record.timestamp_in and latest_record.timestamp_out is None:
                latest_record.timestamp_out = now_utc
                latest_record.status = 'Clocked Out'
                db.session.commit()
                message = f'{employee.name}, you have been clocked OUT.'
                speech_text = f'{employee.name} logged out'
                status_code = 'success'
            else:
                new_attendance = AttendanceRecord(
                    employee_id=employee.id,
                    date=today,
                    timestamp_in=now_utc,
                    timestamp_out=None,
                    status='Clocked In'
                )
                db.session.add(new_attendance)
                db.session.commit()
                message = f'Attendance marked IN for {employee.name} ({employee.employee_id}). Welcome!'
                speech_text = f'{employee.name} logged in'
                status_code = 'success'

            # Announce via TTS in auto mode
            if auto_mode and speech_text:
                speak_async(speech_text)

            return jsonify({'status': status_code, 'message': message, 'employee_id': employee_id_recognized})
        else:
            return jsonify({'status': 'error', 'message': 'Recognized face, but employee not found.'})
    else:
        return jsonify({'status': 'error', 'message': 'Face not recognized or no face detected.'})

# --- Admin Routes (unchanged) ---
@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if current_user.is_authenticated:
        return redirect(url_for('admin_dashboard'))
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        admin = AdminUser.query.filter_by(username=username).first()
        if admin and check_password_hash(admin.password_hash, password):
            login_user(admin)
            flash('Login successful!', 'success')
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Login Unsuccessful. Check username/password.', 'danger')
    return render_template('admin_login.html', title="Admin Login")

@app.route('/admin/logout')
@login_required
def admin_logout():
    logout_user()
    flash('You have been logged out.', 'info')
    return redirect(url_for('admin_login'))

@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    total_employees = Employee.query.count()
    today = date.today()

    present_today_count = db.session.query(AttendanceRecord.employee_id).\
        filter(AttendanceRecord.date == today, AttendanceRecord.timestamp_in.isnot(None)).\
        distinct().count()

    absent_today_count = total_employees - present_today_count

    attendance_summary_today_dict = {
        'Present': present_today_count,
        'Absent': absent_today_count if absent_today_count >= 0 else 0
    }

    labels_last_7_days = []
    counts_last_7_days = []
    for i in range(6, -1, -1):
        current_date_loop = today - timedelta(days=i)
        labels_last_7_days.append(current_date_loop.strftime('%a, %b %d'))
        count = db.session.query(AttendanceRecord.employee_id).\
            filter(AttendanceRecord.date == current_date_loop, AttendanceRecord.timestamp_in.isnot(None)).\
            distinct().count()
        counts_last_7_days.append(count)

    last_7_days_chart_data_dict = {
        'labels': labels_last_7_days,
        'counts': counts_last_7_days
    }

    return render_template('admin_dashboard.html', title="Admin Dashboard",
                           total_employees=total_employees,
                           today_attendance_count=present_today_count,
                           attendance_summary_today=attendance_summary_today_dict,
                           last_7_days_chart_data=last_7_days_chart_data_dict,
                           attendance_summary_today_json=json.dumps(attendance_summary_today_dict),
                           last_7_days_chart_data_json=json.dumps(last_7_days_chart_data_dict)
                           )

@app.route('/admin/add_employee', methods=['GET', 'POST'])
@login_required
def add_employee():
    if request.method == 'POST':
        employee_id = request.form.get('employee_id')
        name = request.form.get('name')
        phone = request.form.get('phone')
        address = request.form.get('address')
        dob_str = request.form.get('dob')
        designation = request.form.get('designation')
        if not all([employee_id, name]):
            flash('Employee ID and Name are required.', 'danger')
            return redirect(request.url)
        if Employee.query.filter_by(employee_id=employee_id).first():
            flash(f'Employee ID {employee_id} already exists.', 'danger')
            return redirect(request.url)
        dob_obj = None
        if dob_str:
            try:
                dob_obj = datetime.strptime(dob_str, '%Y-%m-%d').date()
            except ValueError:
                flash('Invalid DOB format. Use YYYY-MM-DD.', 'danger')
                return redirect(request.url)
        new_employee = Employee(
            employee_id=employee_id, name=name, phone=phone, address=address, dob=dob_obj,
            designation=designation,
            face_encodings=json.dumps([])
        )
        db.session.add(new_employee)
        db.session.commit()
        flash(f'Employee {name} added. Proceed with face training.', 'success')
        return redirect(url_for('train_face_realtime_page', emp_db_id=new_employee.id))
    return render_template('add_employee.html', title="Add New Employee")

@app.route('/admin/manage_employees')
@login_required
def manage_employees():
    employees = Employee.query.order_by(Employee.name).all()
    return render_template('manage_employees.html', title="Manage Employees", employees=employees)

@app.route('/admin/edit_employee/<int:emp_db_id>', methods=['GET', 'POST'])
@login_required
def edit_employee(emp_db_id):
    employee = db.session.get(Employee, emp_db_id)
    if not employee:
        flash('Employee not found.', 'danger')
        return redirect(url_for('manage_employees'))
    if request.method == 'POST':
        employee.name = request.form.get('name', employee.name)
        employee.phone = request.form.get('phone', employee.phone)
        employee.address = request.form.get('address', employee.address)
        employee.designation = request.form.get('designation', employee.designation)
        dob_str = request.form.get('dob')
        if dob_str:
            try:
                employee.dob = datetime.strptime(dob_str, '%Y-%m-%d').date()
            except ValueError:
                flash('Invalid DOB format. Use YYYY-MM-DD.', 'warning')
        images_from_form = request.files.getlist('face_images')
        if images_from_form and images_from_form[0].filename != '':
            new_embeddings_from_files_lists = []
            employee_upload_temp_dir = os.path.join(app.config['KNOWN_FACES_DIR'], employee.employee_id, "temp_edit_uploads")
            if not os.path.exists(employee_upload_temp_dir):
                os.makedirs(employee_upload_temp_dir)
            for idx, image_file_storage in enumerate(images_from_form):
                if image_file_storage.filename == '': continue
                safe_filename = f"edit_upload_face_{idx+1}{os.path.splitext(image_file_storage.filename)[1]}"
                temp_image_path = os.path.join(employee_upload_temp_dir, safe_filename)
                try:
                    image_file_storage.save(temp_image_path)
                    cv2_img_for_detection = cv2.imread(temp_image_path)
                    if cv2_img_for_detection is None:
                        os.remove(temp_image_path)
                        continue
                    _boxes, aligned_tensors = detect_faces_and_get_aligned_tensors(cv2_img_for_detection)
                    if aligned_tensors is not None and len(aligned_tensors) > 0:
                        embedding_tensor = get_face_embedding_from_aligned_tensor(aligned_tensors[0])
                        if embedding_tensor is not None:
                            new_embeddings_from_files_lists.append(embedding_tensor.tolist())
                    os.remove(temp_image_path)
                except Exception as e_upload:
                    print(f"ERROR processing uploaded file {safe_filename}: {e_upload}")
                    if os.path.exists(temp_image_path): os.remove(temp_image_path)
            if os.path.exists(employee_upload_temp_dir) and not os.listdir(employee_upload_temp_dir):
                os.rmdir(employee_upload_temp_dir)
            elif os.path.exists(employee_upload_temp_dir) and os.listdir(employee_upload_temp_dir):
                print(f"WARNING: Temp upload directory not empty after processing: {employee_upload_temp_dir}")

            if new_embeddings_from_files_lists:
                employee.face_encodings = json.dumps(new_embeddings_from_files_lists)
                flash(f"{len(new_embeddings_from_files_lists)} new embeddings generated. Previous replaced.", "info")
                load_known_face_embeddings_from_db_to_gallery()
            else:
                flash("No valid new embeddings from uploaded files. Existing embeddings unchanged.", "warning")
        db.session.commit()
        flash(f'Employee {employee.name} updated.', 'success')
        return redirect(url_for('manage_employees'))
    return render_template('edit_employee.html', title=f"Edit {employee.name}", employee=employee)

@app.route('/admin/delete_employee/<int:emp_db_id>', methods=['POST'])
@login_required
def delete_employee(emp_db_id):
    employee = db.session.get(Employee, emp_db_id)
    if employee:
        try:
            employee_data_dir = os.path.join(app.config['KNOWN_FACES_DIR'], employee.employee_id)
            if os.path.exists(employee_data_dir):
                try:
                    shutil.rmtree(employee_data_dir)
                    print(f"INFO: Removed data directory for employee {employee.employee_id}: {employee_data_dir}")
                except Exception as e_rm:
                    print(f"ERROR: Could not remove directory {employee_data_dir}: {e_rm}")
            db.session.delete(employee)
            db.session.commit()
            load_known_face_embeddings_from_db_to_gallery()
            flash(f'Employee {employee.name} and their records deleted.', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error deleting employee: {str(e)}', 'danger')
    else:
        flash('Employee not found.', 'danger')
    return redirect(url_for('manage_employees'))

@app.route('/admin/train_face_realtime/<int:emp_db_id>')
@login_required
def train_face_realtime_page(emp_db_id):
    employee = db.session.get(Employee, emp_db_id)
    if not employee:
        flash('Employee not found.', 'danger')
        return redirect(url_for('manage_employees'))
    num_existing_encodings = 0
    if employee.face_encodings and employee.face_encodings != '[]':
        try:
            existing_encodings = json.loads(employee.face_encodings)
            num_existing_encodings = len(existing_encodings)
        except json.JSONDecodeError:
            num_existing_encodings = 0
    return render_template('train_employee_face_realtime.html',
                           title=f"Train Face: {employee.name}",
                           employee=employee,
                           num_existing_encodings=num_existing_encodings)

@app.route('/admin/capture_training_frame/<int:emp_db_id>', methods=['POST'])
@login_required
def capture_training_frame(emp_db_id):
    employee = db.session.get(Employee, emp_db_id)
    if not employee: return jsonify({'status': 'error', 'message': 'Employee not found.'}), 404
    if mtcnn_detector is None or face_embedder_new is None:
        return jsonify({'status': 'error', 'message': 'Face models not ready.'}), 503
    data = request.get_json()
    if not data or 'image_data' not in data:
        return jsonify({'status': 'error', 'message': 'No image data.'}), 400
    image_data_url = data['image_data']
    try:
        header, encoded = image_data_url.split(",", 1)
        image_data_bytes = base64.b64decode(encoded)
        nparr = np.frombuffer(image_data_bytes, np.uint8)
        cv2_frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'Error decoding image: {str(e)}'}), 400
    if cv2_frame is None: return jsonify({'status': 'error', 'message': 'Could not decode image.'}), 400
    _face_boxes, aligned_face_tensors = detect_faces_and_get_aligned_tensors(cv2_frame)
    if aligned_face_tensors is None or len(aligned_face_tensors) == 0:
        return jsonify({'status': 'error', 'message': 'No face detected.'})
    if len(aligned_face_tensors) > 1:
        return jsonify({'status': 'warning', 'message': f'Multiple ({len(aligned_face_tensors)}) faces detected.'})
    new_embedding_tensor = get_face_embedding_from_aligned_tensor(aligned_face_tensors[0])
    if new_embedding_tensor is None:
        return jsonify({'status': 'error', 'message': 'Could not generate embedding.'})
    new_embedding_list = new_embedding_tensor.tolist()
    try:
        current_embeddings_list_of_lists = json.loads(employee.face_encodings or '[]')
    except json.JSONDecodeError:
        current_embeddings_list_of_lists = []
    current_embeddings_list_of_lists.append(new_embedding_list)
    employee.face_encodings = json.dumps(current_embeddings_list_of_lists)
    db.session.commit()
    load_known_face_embeddings_from_db_to_gallery()
    return jsonify({
        'status': 'success',
        'message': f'Frame captured for {employee.name}.',
        'total_encodings_for_employee': len(current_embeddings_list_of_lists)
    })

@app.route('/admin/view_attendance', methods=['GET', 'POST'])
@login_required
def view_attendance():
    selected_date_str = request.form.get('attendance_date') or request.args.get('attendance_date') or date.today().strftime('%Y-%m-%d')

    try:
        selected_date_obj = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
    except ValueError:
        flash("Invalid date format. Showing today's attendance.", "warning")
        selected_date_obj = date.today()
        selected_date_str = selected_date_obj.strftime('%Y-%m-%d')

    attendance_records = db.session.query(AttendanceRecord, Employee.name, Employee.employee_id, Employee.designation).\
        join(Employee, AttendanceRecord.employee_id == Employee.id).\
        filter(AttendanceRecord.date == selected_date_obj).\
        order_by(Employee.employee_id, AttendanceRecord.timestamp_in).\
        all()
    return render_template('view_attendance.html', title="View Attendance",
                           records_with_employee_info=attendance_records,
                           selected_date=selected_date_str)

@app.route('/admin/export_attendance', methods=['POST'])
@login_required
def export_attendance():
    start_date_str = request.form.get('start_date')
    end_date_str = request.form.get('end_date')

    if not start_date_str or not end_date_str:
        flash("Please select both start and end dates for export.", "danger")
        return redirect(url_for('view_attendance', attendance_date=date.today().strftime('%Y-%m-%d')))

    try:
        start_date_obj = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        flash("Invalid date format for export. Please use YYYY-MM-DD.", "danger")
        return redirect(url_for('view_attendance', attendance_date=end_date_str))

    if end_date_obj < start_date_obj:
        flash("End date cannot be before start date for export.", "danger")
        return redirect(url_for('view_attendance', attendance_date=end_date_str))

    records_to_export = db.session.query(AttendanceRecord, Employee.name, Employee.employee_id, Employee.designation).\
        join(Employee, AttendanceRecord.employee_id == Employee.id).\
        filter(AttendanceRecord.date >= start_date_obj, AttendanceRecord.date <= end_date_obj).\
        order_by(AttendanceRecord.date, Employee.employee_id, AttendanceRecord.timestamp_in).\
        all()

    if not records_to_export:
        flash("No attendance records found for the selected date range to export.", "info")
        return redirect(url_for('view_attendance', attendance_date=end_date_str))

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    headers = ["Date", "Employee ID", "Employee Name", "Designation", "Time In", "Time Out", "Duration (HH:MM:SS)", "Status"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for record_obj, emp_name_str, emp_id_val, emp_designation_val in records_to_export:
        time_in_str = record_obj.timestamp_in.strftime("%H:%M:%S") if record_obj.timestamp_in else ""
        time_out_str = record_obj.timestamp_out.strftime("%H:%M:%S") if record_obj.timestamp_out else ""
        duration_str = ""
        if record_obj.timestamp_in and record_obj.timestamp_out:
            duration = record_obj.timestamp_out - record_obj.timestamp_in
            hours, remainder = divmod(duration.total_seconds(), 3600)
            minutes, seconds = divmod(remainder, 60)
            duration_str = f"{int(hours):02}:{int(minutes):02}:{int(seconds):02}"

        ws.append([
            record_obj.date.strftime("%Y-%m-%d"),
            emp_id_val,
            emp_name_str,
            emp_designation_val or "",
            time_in_str,
            time_out_str,
            duration_str,
            record_obj.status
        ])

    for col_idx, column_cells in enumerate(ws.columns):
        max_length = 0
        column = openpyxl.utils.get_column_letter(col_idx + 1)
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name=f'attendance_report_{start_date_str}_to_{end_date_str}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/admin/edit_attendance/<int:att_id>', methods=['GET', 'POST'])
@login_required
def edit_attendance(att_id):
    attendance_record = db.session.get(AttendanceRecord, att_id)
    if not attendance_record:
        flash("Attendance record not found.", "danger")
        return redirect(url_for('view_attendance'))
    employee = db.session.get(Employee, attendance_record.employee_id)
    if not employee:
        flash("Associated employee not found.", "danger")
        return redirect(url_for('view_attendance'))
    if request.method == 'POST':
        new_status = request.form.get('status')
        time_in_str = request.form.get('time_in')
        time_out_str = request.form.get('time_out')
        attendance_record.status = new_status
        if time_in_str:
            try:
                time_format = '%H:%M:%S' if len(time_in_str.split(':')) == 3 else '%H:%M'
                time_in_obj = datetime.strptime(time_in_str, time_format).time()
                attendance_record.timestamp_in = datetime.combine(attendance_record.date, time_in_obj)
            except ValueError: flash(f"Invalid Time In: '{time_in_str}'. Use HH:MM or HH:MM:SS.", "warning")
        else:
            attendance_record.timestamp_in = None
        if time_out_str:
            try:
                time_format = '%H:%M:%S' if len(time_out_str.split(':')) == 3 else '%H:%M'
                time_out_obj = datetime.strptime(time_out_str, time_format).time()
                attendance_record.timestamp_out = datetime.combine(attendance_record.date, time_out_obj)
            except ValueError: flash(f"Invalid Time Out: '{time_out_str}'. Use HH:MM or HH:MM:SS.", "warning")
        else:
            attendance_record.timestamp_out = None
        db.session.commit()
        flash("Attendance record updated.", "success")
        return redirect(url_for('view_attendance', attendance_date=attendance_record.date.strftime('%Y-%m-%d')))
    return render_template('edit_attendance.html', title="Edit Attendance", record=attendance_record, employee_name=employee.name)

def create_initial_admin():
    with app.app_context():
        if not AdminUser.query.first():
            username = os.environ.get('ADMIN_USERNAME', "admin")
            password = os.environ.get('ADMIN_PASSWORD', "password123")
            hashed_password = generate_password_hash(password)
            admin = AdminUser(username=username, password_hash=hashed_password)
            db.session.add(admin)
            db.session.commit()
            print(f"INFO: Admin user '{username}' created. PLEASE CHANGE DEFAULT PASSWORD.")

# --- App Initialization Block ---
with app.app_context():
    import openpyxl
    db.create_all()
    create_initial_admin()
    initialize_tts()
    models_loaded_successfully = load_new_face_models()
    if models_loaded_successfully:
        load_known_face_embeddings_from_db_to_gallery()
    else:
        print("CRITICAL: Face models failed to load. System may not function correctly.")

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)